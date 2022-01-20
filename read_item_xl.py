from inspect import CO_COROUTINE
from msilib.sequence import tables
from multiprocessing import cpu_count
import re


def txt2num( q, t):
    if q==None :
       return None
    elif isinstance(q, int) or isinstance(q, float):
       return q
    else:
       r_=re.findall('[0-9.]+',q)
       if len(r_) == 0:
           return None
       elif t=="int":
           return int (r_[0])           
       else:
           return float(r_[0])
        


from openpyxl import Workbook  

out_wb = Workbook()
out_ws = out_wb.active
source_file_name="doc/out_asset_data.xlsx"
out_file_name="doc/out_item.xlsx"

from openpyxl import load_workbook   

field_colum_names=  ['物品編號', '年度', '憑單編號','傳票號碼', '物品', '牌子及型號', 'S/N (P/N)', '數量', '單價(M$) ', '總值(M$)', '攤折淨值(M$)', '存放地點', '資助', '資助項目名稱', '備註', '供應商', '登賬日期','cate_','unit']
ignore_ws=['資產分類表','工作表1']
ignore_col=['異動   原因','異動原因','*    攤折完成',  '*  攤折完', '* 攤折完成','*  攤折完成','^ 投保產物', '投保項備註','學校資金']
catetype={}
out_ws.append(field_colum_names)

wb = load_workbook(filename = source_file_name)
for sn_ in wb.sheetnames:
    if sn_ != "Sheet" :continue    
    ws=wb[sn_]
    for ridx in range(1,5000):
        row_ =[]
        for cidx in range(65,88):
            colum_name_=ws[f"{chr(cidx)}{ridx}"].value
            #if colum_name_==None: break
            row_.append(colum_name_)
        year_=row_[1]
        if year_==None: break;
        CATNO_=row_[17]
        #if CATNO_!="4271": continue

        year_=year_.replace("/","")
        rlen_=len(row_)
        catetype_=str(row_[rlen_-1]).replace("999","000")
        row_[rlen_-1]=catetype_
        year_=row_[1].replace("/","")
        catekey_=f"{year_}{catetype_}"
        row_.append(catekey_)
        Get_id=catetype.get(catekey_)
        if Get_id==None : 
            Get_id=0
        
        q_=txt2num(row_[7],"int")
        p_=txt2num(row_[8],"float")
        t_=txt2num(row_[9],"float")
        fundm_=row_[12]
        cc=str(row_[2])[2:6]
        if (q_!=None and q_>1) and (p_!=None and p_>1000) and row_[1] in ["16/17","17/18","18/19","19/20","20/21"] and cc in ["4271","4273","4274","4280","4295","4296"]:
            for i in range(0,q_):
                if i>0: row_[12]=0
                catetype[catekey_]=Get_id+1+i
                print([*row_,f"{catekey_}{str(Get_id+1+i).zfill(3)}",1,p_,q_])
                out_ws.append([*row_,f"{catekey_}{str(Get_id+1+i).zfill(3)}",1,p_,q_])
        else:
            catetype[catekey_]=Get_id+1
            row_.append(f"{catekey_}{str(Get_id+1).zfill(3)}")
            row_.append(q_)
            row_.append(p_)
            row_.append(t_)
            print(row_)
            out_ws.append(row_)
        
        

out_wb.save(out_file_name)