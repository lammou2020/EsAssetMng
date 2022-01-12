from openpyxl import Workbook  
out_wb = Workbook()
out_ws = out_wb.active
# add a simple formula
# ws["A1"] = "=SUM(1, 1)"

from openpyxl import load_workbook   

field_colum_names=  ['物品編號', '年度', '憑單編號',            '傳票號碼', '物品', '牌子及型號', 'S/N (P/N)', '數量', '單價(M$) ', '總值(M$)', '攤折淨值(M$)', '存放地點', '資助', '資助項目名稱', '備註', '供應商', '登賬日期','cate_','unit']
ignore_ws=['資產分類表','工作表1']
ignore_col=['異動   原因','異動原因','*    攤折完成',  '*  攤折完', '* 攤折完成','*  攤折完成','^ 投保產物', '投保項備註','學校資金']

out_ws.append(field_colum_names)

wb = load_workbook(filename = 'doc/asset_data.xlsx')
for sn_ in wb.sheetnames:
    if sn_ in ignore_ws:continue    
    ws=wb[sn_]
    col_f_n_ =[]
    col_i_n_ =[]
    f_col_=None
    for idx in range(65,90):
        colum_name_=ws[f"{chr(idx)}{6}"].value
        if colum_name_==None: break
        if colum_name_ in ignore_col: continue        
        if colum_name_=="年度":f_col_=chr(idx)
        if f_col_==None and colum_name_=="憑單編號":f_col_=chr(idx)
        col_f_n_.append(colum_name_)
        col_i_n_.append(chr(idx))

    for f_ in col_f_n_:
        print(field_colum_names.index(f_))

    for ridx in range(7,750):
        key_v=ws[f"{f_col_}{ridx}"].value
        if key_v==None or key_v=="" : break
        row_=['', '', '', '', '', '', '', '', '', '','', '', '', '', '', '', '',sn_]
        for idx,v_ in enumerate(col_f_n_):
            cell_v=ws[f"{col_i_n_[idx]}{ridx}"].value
            row_[field_colum_names.index(v_)]=cell_v
        out_ws.append(row_)

out_wb.save("doc/out_asset_data.xlsx")